import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By

import re

import os


def create_file(columns, columns_codes, file_name):
    """
    Создание .xlsx файла
    :param columns: названия столбцов
    :param columns_codes: ячейки
    :param file_name: название файла с расширением
    :return:
    """
    if os.path.exists(file_name):
        print("Файл уже существует, запись продолжится")

        workbook = openpyxl.load_workbook('viking_result.xlsx')

        # выбираем активный лист
        sheet = workbook.active

        # задаем номер столбца, который нужно проверить
        column_number = 1

        # считаем количество заполненных ячеек в столбце
        filled_cells = 0
        for cell in sheet['A']:
            if cell.value is not None:
                filled_cells += 1

        filled_cells -= 1  # не считаем первую ячейку - название столбца

        pages_worked = filled_cells // 48
        print(pages_worked)

        return pages_worked
    else:
        print("Файл не существует")
        workbook = openpyxl.Workbook()  # создание результирующего файла
        sheet = workbook.active

        for i in range(len(columns)):
            sheet[f'{columns_codes[i]}1'] = columns[i]

        workbook.save(file_name)

        return 0


def create_driver(url):
    """
    Функция создает драйвер для парсинга
    :param url: адрес страницы
    :return: драйвер
    """

    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_argument("--headless")  # работа без открытия браузера
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/105.0.0.0 Safari/537.36")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()

    driver.get(url)

    return driver


def load_to_excel(pathfile, data):
    """
    Функция добавляет строчку в .xlsx-файл
    :param pathfile: путь к файлу
    :param data: строчка
    :return:
    """

    workbook = openpyxl.load_workbook(pathfile)
    sheet = workbook.active
    for i in data:
        sheet.append(i)

    workbook.save(pathfile)

    return True
